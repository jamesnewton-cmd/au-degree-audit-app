import unittest
from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

import scheduling_main as sched


def _auth_headers(token: str) -> dict[str, str]:
    return {"Authorization": f"Bearer {token}"}


def _issue_auth_token(client: TestClient, email: str = "advisor@anderson.edu") -> str:
    req = client.post("/auth/request-code", json={"email": email})
    if req.status_code != 200:
        raise AssertionError(f"Unable to request auth code: {req.status_code} {req.text}")
    verify = client.post("/auth/verify-code", json={"email": email, "code": sched.AUTH_STATIC_TEST_CODE})
    if verify.status_code != 200:
        raise AssertionError(f"Unable to verify auth code: {verify.status_code} {verify.text}")
    token = verify.json().get("access_token")
    if not token:
        raise AssertionError("Missing access token in auth response.")
    return token


def _build_class_listings_xlsx() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Class Listings"
    ws.append(
        [
            "Course Code",
            "Course Title",
            "Section",
            "Instructor",
            "Capacity",
            "Days",
            "Start Time",
            "End Time",
        ]
    )
    ws.append(["MATH 1010", "College Algebra", "A", "Prof Euler", 1, "MWF", "09:00 AM", "10:00 AM"])
    ws.append(["MATH 1010", "College Algebra", "B", "Prof Noether", 1, "TR", "11:00 AM", "12:15 PM"])
    ws.append(["ENGL 1010", "Composition I", "A", "Prof Angelou", 2, "MWF", "09:30 AM", "10:20 AM"])
    ws.append(["HIST 1000", "World History", "A", "Prof Herodotus", 1, "TR", "11:00 AM", "12:15 PM"])
    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def _build_student_requests_xlsx() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Requests"
    ws.append(["Student ID", "Student Name", "Advisor", "Request 1", "Request 2", "Request 3"])
    ws.append(["S001", "Alice Smith", "Dr. Rivera", "MATH 1010", "ENGL 1010", "HIST 1000"])
    ws.append(["S002", "Bob Jones", "Dr. Rivera", "MATH 1010", "HIST 1000", "ENGL 1010"])
    ws.append(["S003", "Cara Lee", "Dr. Ahmed", "MATH 1010", "ENGL 1010", "BIOL 1000"])
    out = BytesIO()
    wb.save(out)
    return out.getvalue()


class SchedulingAppTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls._orig_auth_static = sched.AUTH_STATIC_TEST_CODE
        sched.AUTH_STATIC_TEST_CODE = "123456"
        sched.AUTH_CODES.clear()
        sched.AUTH_SESSIONS.clear()
        cls.client = TestClient(sched.app)
        cls.headers = _auth_headers(_issue_auth_token(cls.client))

    @classmethod
    def tearDownClass(cls):
        sched.AUTH_STATIC_TEST_CODE = cls._orig_auth_static
        sched.AUTH_CODES.clear()
        sched.AUTH_SESSIONS.clear()

    def test_root_requires_auth(self):
        self.assertEqual(self.client.get("/").status_code, 401)
        authed = self.client.get("/", headers=self.headers)
        self.assertEqual(authed.status_code, 200)
        self.assertIn("Professor / Advisor Scheduling", authed.text)

    def test_generate_returns_xlsx_and_summary_headers(self):
        files = {
            "class_listings": (
                "class_listings.xlsx",
                _build_class_listings_xlsx(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
            "student_requests": (
                "student_requests.xlsx",
                _build_student_requests_xlsx(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
        }
        data = {"max_courses_per_student": "3"}
        response = self.client.post("/generate", data=data, files=files, headers=self.headers)
        self.assertEqual(response.status_code, 200)
        self.assertIn(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            response.headers.get("content-type", ""),
        )
        self.assertEqual(response.headers.get("x-scheduling-students"), "3")
        self.assertEqual(response.headers.get("x-scheduling-assignments"), "5")
        self.assertEqual(response.headers.get("x-scheduling-unscheduled"), "4")

        workbook = load_workbook(BytesIO(response.content), data_only=True)
        self.assertIn("Student Schedules", workbook.sheetnames)
        self.assertIn("Unscheduled Requests", workbook.sheetnames)
        self.assertIn("Section Fill", workbook.sheetnames)
        self.assertIn("Summary", workbook.sheetnames)

    def test_generate_rejects_invalid_max_courses(self):
        files = {
            "class_listings": (
                "class_listings.xlsx",
                _build_class_listings_xlsx(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
            "student_requests": (
                "student_requests.xlsx",
                _build_student_requests_xlsx(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
        }
        response = self.client.post(
            "/generate",
            data={"max_courses_per_student": "0"},
            files=files,
            headers=self.headers,
        )
        self.assertEqual(response.status_code, 400)
        self.assertIn("at least 1", response.text)


if __name__ == "__main__":
    unittest.main()
