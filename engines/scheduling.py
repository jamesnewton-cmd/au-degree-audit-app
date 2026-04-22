"""
Scheduling utilities for advisor/professor planning from Excel uploads.

This module provides:
- Excel parsing for class listings and student course requests
- Conflict-aware schedule generation with seat capacity limits
- Excel export of generated student schedules and diagnostics
"""

from __future__ import annotations

import datetime as dt
import io
import re
from dataclasses import dataclass, field
from typing import Any

from openpyxl import Workbook, load_workbook


DAY_ORDER = ("M", "T", "W", "R", "F", "S", "U")


@dataclass
class MeetingBlock:
    days: set[str]
    start_minute: int
    end_minute: int


@dataclass
class ClassSection:
    section_id: str
    course_code: str
    course_title: str
    section: str
    instructor: str
    capacity: int
    remaining: int
    meetings: list[MeetingBlock] = field(default_factory=list)


@dataclass
class StudentRequest:
    student_id: str
    student_name: str
    advisor: str
    requests: list[str]


def _normalize_header(value: Any) -> str:
    raw = str(value or "").strip().lower()
    raw = re.sub(r"[^a-z0-9]+", "_", raw)
    return raw.strip("_")


def _text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _norm_course_code(value: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9]+", "_", (value or "").strip().upper())
    return re.sub(r"_+", "_", cleaned).strip("_")


def _display_course_code(value: str) -> str:
    return (value or "").replace("_", " ")


def _parse_capacity(value: Any, default: int = 30) -> int:
    if value is None or str(value).strip() == "":
        return default
    try:
        parsed = int(float(value))
        return max(1, parsed)
    except (TypeError, ValueError):
        return default


def _parse_time_to_minutes(value: Any) -> int | None:
    if value is None:
        return None

    if isinstance(value, dt.datetime):
        return value.hour * 60 + value.minute
    if isinstance(value, dt.time):
        return value.hour * 60 + value.minute

    if isinstance(value, (int, float)):
        number = float(value)
        # Excel time serial fraction (0.0 - 1.0)
        if 0 <= number <= 1:
            return int(round(number * 24 * 60))
        # 900 or 1330 patterns
        as_int = int(number)
        if 0 <= as_int <= 2359 and as_int % 100 < 60:
            return (as_int // 100) * 60 + (as_int % 100)

    raw = str(value).strip().upper().replace(".", "")
    if not raw:
        return None

    for fmt in ("%H:%M", "%I:%M%p", "%I%p", "%H%M", "%I:%M %p"):
        try:
            parsed = dt.datetime.strptime(raw.replace(" ", ""), fmt)
            return parsed.hour * 60 + parsed.minute
        except ValueError:
            continue

    # Fallback for values like "9:00 AM"
    try:
        parsed = dt.datetime.strptime(raw, "%I:%M %p")
        return parsed.hour * 60 + parsed.minute
    except ValueError:
        return None


def _parse_days(value: Any) -> set[str]:
    raw = str(value or "").strip().upper()
    if not raw:
        return set()

    if "ASYNC" in raw or "ONLINE" in raw:
        return set()

    normalized = (
        raw.replace("TH", "R")
        .replace("TR", "TR")
        .replace("TU", "T")
        .replace("TUES", "T")
        .replace("MON", "M")
        .replace("WED", "W")
        .replace("THU", "R")
        .replace("FRI", "F")
        .replace("SAT", "S")
        .replace("SUN", "U")
    )
    normalized = re.sub(r"[^MTWRFSU]", "", normalized)
    return {ch for ch in normalized if ch in DAY_ORDER}


def _format_days(days: set[str]) -> str:
    if not days:
        return "Asynchronous / TBA"
    return "".join(ch for ch in DAY_ORDER if ch in days)


def _format_minutes(minutes: int | None) -> str:
    if minutes is None:
        return "TBA"
    hours = minutes // 60
    mins = minutes % 60
    suffix = "AM" if hours < 12 else "PM"
    display_h = hours % 12
    if display_h == 0:
        display_h = 12
    return f"{display_h}:{mins:02d} {suffix}"


def _extract_meetings(days: Any, start: Any, end: Any, meeting_time: Any) -> list[MeetingBlock]:
    blocks: list[MeetingBlock] = []

    day_set = _parse_days(days)
    start_minute = _parse_time_to_minutes(start)
    end_minute = _parse_time_to_minutes(end)
    if day_set and start_minute is not None and end_minute is not None and end_minute > start_minute:
        blocks.append(MeetingBlock(days=day_set, start_minute=start_minute, end_minute=end_minute))
        return blocks

    raw = str(meeting_time or "").strip()
    if not raw:
        return blocks

    parts = [p.strip() for p in re.split(r"[;|]+", raw) if p.strip()]
    pattern = re.compile(
        r"([A-Za-z/,&\s]+)\s+([0-9: ]+[APMapm]{0,2})\s*[-–]\s*([0-9: ]+[APMapm]{0,2})"
    )
    for part in parts:
        match = pattern.search(part)
        if not match:
            continue
        part_days = _parse_days(match.group(1))
        part_start = _parse_time_to_minutes(match.group(2))
        part_end = _parse_time_to_minutes(match.group(3))
        if part_days and part_start is not None and part_end is not None and part_end > part_start:
            blocks.append(
                MeetingBlock(days=part_days, start_minute=part_start, end_minute=part_end)
            )
    return blocks


def _column_map(headers: list[str]) -> dict[str, int]:
    return {name: idx for idx, name in enumerate(headers)}


def _find_first(col_map: dict[str, int], aliases: tuple[str, ...]) -> int | None:
    for alias in aliases:
        if alias in col_map:
            return col_map[alias]
    return None


def parse_class_sections_workbook(blob: bytes) -> tuple[list[ClassSection], list[str]]:
    workbook = load_workbook(io.BytesIO(blob), data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Class listing workbook is empty.")

    headers = [_normalize_header(v) for v in rows[0]]
    col_map = _column_map(headers)

    code_idx = _find_first(col_map, ("course_code", "course", "class_code", "crn_course"))
    if code_idx is None:
        raise ValueError(
            "Class listing is missing required column: Course Code (or equivalent)."
        )

    section_idx = _find_first(col_map, ("section", "section_number", "class_section"))
    title_idx = _find_first(col_map, ("course_title", "title", "class_title", "name"))
    instructor_idx = _find_first(col_map, ("instructor", "professor", "faculty", "advisor"))
    capacity_idx = _find_first(
        col_map, ("capacity", "seats", "max_enrollment", "enrollment_cap", "max_size")
    )
    days_idx = _find_first(col_map, ("days", "meeting_days", "day_pattern"))
    start_idx = _find_first(col_map, ("start_time", "start", "time_start"))
    end_idx = _find_first(col_map, ("end_time", "end", "time_end"))
    meeting_idx = _find_first(col_map, ("meeting_time", "meeting_pattern", "time"))

    warnings: list[str] = []
    sections: list[ClassSection] = []

    for i, row in enumerate(rows[1:], start=2):
        code_raw = _text(row[code_idx] if code_idx < len(row) else "")
        if not code_raw:
            continue

        course_code = _norm_course_code(code_raw)
        if not course_code:
            warnings.append(f"Row {i}: invalid course code '{code_raw}'.")
            continue

        section = _text(row[section_idx] if section_idx is not None and section_idx < len(row) else "")
        section = section or "A"
        title = _text(row[title_idx] if title_idx is not None and title_idx < len(row) else "")
        instructor = _text(
            row[instructor_idx] if instructor_idx is not None and instructor_idx < len(row) else ""
        )
        capacity = _parse_capacity(
            row[capacity_idx] if capacity_idx is not None and capacity_idx < len(row) else None
        )

        meetings = _extract_meetings(
            row[days_idx] if days_idx is not None and days_idx < len(row) else None,
            row[start_idx] if start_idx is not None and start_idx < len(row) else None,
            row[end_idx] if end_idx is not None and end_idx < len(row) else None,
            row[meeting_idx] if meeting_idx is not None and meeting_idx < len(row) else None,
        )

        section_id = f"{course_code}::{section}::{i}"
        sections.append(
            ClassSection(
                section_id=section_id,
                course_code=course_code,
                course_title=title,
                section=section,
                instructor=instructor,
                capacity=capacity,
                remaining=capacity,
                meetings=meetings,
            )
        )

    if not sections:
        raise ValueError("No class sections were found in the class listing workbook.")
    return sections, warnings


def parse_student_requests_workbook(blob: bytes) -> tuple[list[StudentRequest], list[str]]:
    workbook = load_workbook(io.BytesIO(blob), data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Student request workbook is empty.")

    headers = [_normalize_header(v) for v in rows[0]]
    col_map = _column_map(headers)
    id_idx = _find_first(col_map, ("student_id", "id", "student_number"))
    name_idx = _find_first(col_map, ("student_name", "name", "student"))
    advisor_idx = _find_first(col_map, ("advisor", "advisor_name", "faculty_advisor"))
    list_idx = _find_first(col_map, ("requested_courses", "course_requests", "requests"))

    request_cols = []
    for idx, header in enumerate(headers):
        if re.match(r"request_?\d+$", header):
            request_cols.append(idx)
        elif re.match(r"course_?\d+$", header):
            request_cols.append(idx)
        elif header.startswith("choice_"):
            request_cols.append(idx)
    request_cols = sorted(set(request_cols))

    if list_idx is None and not request_cols:
        raise ValueError(
            "Student request workbook must include request columns (Request 1, Request 2, ...)."
        )

    warnings: list[str] = []
    students: list[StudentRequest] = []
    for i, row in enumerate(rows[1:], start=2):
        raw_id = _text(row[id_idx] if id_idx is not None and id_idx < len(row) else "")
        raw_name = _text(row[name_idx] if name_idx is not None and name_idx < len(row) else "")
        advisor = _text(row[advisor_idx] if advisor_idx is not None and advisor_idx < len(row) else "")

        student_id = raw_id or f"ROW{i}"
        student_name = raw_name or f"Student {student_id}"

        values: list[str] = []
        if list_idx is not None and list_idx < len(row):
            values.extend(re.split(r"[;,|]+", _text(row[list_idx])))
        for col in request_cols:
            if col < len(row):
                values.append(_text(row[col]))

        requests: list[str] = []
        seen = set()
        for val in values:
            code = _norm_course_code(val)
            if not code or code in seen:
                continue
            seen.add(code)
            requests.append(code)

        if not requests:
            warnings.append(f"Row {i}: no valid course requests for {student_name}.")
            continue

        students.append(
            StudentRequest(
                student_id=student_id,
                student_name=student_name,
                advisor=advisor,
                requests=requests,
            )
        )

    if not students:
        raise ValueError("No student requests were found in the student request workbook.")
    return students, warnings


def _sections_conflict(a: ClassSection, b: ClassSection) -> bool:
    # If either section has no meeting data, treat as no direct conflict.
    if not a.meetings or not b.meetings:
        return False
    for left in a.meetings:
        for right in b.meetings:
            if not (left.days & right.days):
                continue
            if left.start_minute < right.end_minute and right.start_minute < left.end_minute:
                return True
    return False


def _pick_section(candidates: list[ClassSection], assigned: list[ClassSection]) -> tuple[ClassSection | None, str]:
    best: ClassSection | None = None
    saw_open_seat = False
    for section in sorted(candidates, key=lambda s: (-s.remaining, s.section, s.section_id)):
        if section.remaining <= 0:
            continue
        saw_open_seat = True
        if any(_sections_conflict(section, taken) for taken in assigned):
            continue
        best = section
        break
    if best:
        return best, ""
    if not saw_open_seat:
        return None, "No seats available"
    return None, "Time conflict with existing assignment"


def generate_student_schedules(
    sections: list[ClassSection],
    students: list[StudentRequest],
    max_courses_per_student: int = 5,
) -> dict[str, Any]:
    if max_courses_per_student < 1:
        raise ValueError("max_courses_per_student must be at least 1.")

    sections_by_course: dict[str, list[ClassSection]] = {}
    for section in sections:
        sections_by_course.setdefault(section.course_code, []).append(section)

    ordered_students = sorted(students, key=lambda s: (s.advisor.lower(), s.student_name.lower(), s.student_id))
    assigned: dict[str, list[ClassSection]] = {student.student_id: [] for student in ordered_students}
    unscheduled_rows: list[dict[str, Any]] = []

    max_requested = max((len(student.requests) for student in ordered_students), default=0)
    for rank in range(max_requested):
        for student in ordered_students:
            student_assigned = assigned[student.student_id]
            if len(student_assigned) >= max_courses_per_student:
                continue
            if rank >= len(student.requests):
                continue

            wanted_code = student.requests[rank]
            if any(section.course_code == wanted_code for section in student_assigned):
                continue

            candidates = sections_by_course.get(wanted_code, [])
            if not candidates:
                unscheduled_rows.append(
                    {
                        "student_id": student.student_id,
                        "student_name": student.student_name,
                        "advisor": student.advisor,
                        "requested_course": wanted_code,
                        "priority": rank + 1,
                        "reason": "Course not found in class listings",
                    }
                )
                continue

            selected, reason = _pick_section(candidates, student_assigned)
            if selected is None:
                unscheduled_rows.append(
                    {
                        "student_id": student.student_id,
                        "student_name": student.student_name,
                        "advisor": student.advisor,
                        "requested_course": wanted_code,
                        "priority": rank + 1,
                        "reason": reason,
                    }
                )
                continue

            selected.remaining -= 1
            student_assigned.append(selected)

    student_rows: list[dict[str, Any]] = []
    for student in ordered_students:
        for section in sorted(assigned[student.student_id], key=lambda s: (_display_course_code(s.course_code), s.section)):
            if section.meetings:
                days = " / ".join(_format_days(block.days) for block in section.meetings)
                time_range = " / ".join(
                    f"{_format_minutes(block.start_minute)}-{_format_minutes(block.end_minute)}"
                    for block in section.meetings
                )
            else:
                days = "Asynchronous / TBA"
                time_range = "TBA"
            student_rows.append(
                {
                    "student_id": student.student_id,
                    "student_name": student.student_name,
                    "advisor": student.advisor,
                    "course_code": _display_course_code(section.course_code),
                    "course_title": section.course_title,
                    "section": section.section,
                    "instructor": section.instructor,
                    "days": days,
                    "time": time_range,
                }
            )

    section_rows: list[dict[str, Any]] = []
    for section in sorted(sections, key=lambda s: (_display_course_code(s.course_code), s.section)):
        assigned_count = section.capacity - section.remaining
        section_rows.append(
            {
                "course_code": _display_course_code(section.course_code),
                "course_title": section.course_title,
                "section": section.section,
                "instructor": section.instructor,
                "capacity": section.capacity,
                "assigned": assigned_count,
                "remaining": section.remaining,
            }
        )

    return {
        "student_rows": student_rows,
        "unscheduled_rows": unscheduled_rows,
        "section_rows": section_rows,
        "summary": {
            "students": len(ordered_students),
            "sections": len(sections),
            "assignments": len(student_rows),
            "unscheduled": len(unscheduled_rows),
        },
    }


def build_schedule_workbook(result: dict[str, Any], warnings: list[str] | None = None) -> bytes:
    workbook = Workbook()
    student_sheet = workbook.active
    student_sheet.title = "Student Schedules"
    student_sheet.append(
        [
            "Student ID",
            "Student Name",
            "Advisor",
            "Course Code",
            "Course Title",
            "Section",
            "Instructor",
            "Days",
            "Time",
        ]
    )
    for row in result.get("student_rows", []):
        student_sheet.append(
            [
                row.get("student_id", ""),
                row.get("student_name", ""),
                row.get("advisor", ""),
                row.get("course_code", ""),
                row.get("course_title", ""),
                row.get("section", ""),
                row.get("instructor", ""),
                row.get("days", ""),
                row.get("time", ""),
            ]
        )

    unscheduled_sheet = workbook.create_sheet("Unscheduled Requests")
    unscheduled_sheet.append(
        ["Student ID", "Student Name", "Advisor", "Requested Course", "Priority", "Reason"]
    )
    for row in result.get("unscheduled_rows", []):
        unscheduled_sheet.append(
            [
                row.get("student_id", ""),
                row.get("student_name", ""),
                row.get("advisor", ""),
                _display_course_code(row.get("requested_course", "")),
                row.get("priority", ""),
                row.get("reason", ""),
            ]
        )

    section_sheet = workbook.create_sheet("Section Fill")
    section_sheet.append(
        ["Course Code", "Course Title", "Section", "Instructor", "Capacity", "Assigned", "Remaining"]
    )
    for row in result.get("section_rows", []):
        section_sheet.append(
            [
                row.get("course_code", ""),
                row.get("course_title", ""),
                row.get("section", ""),
                row.get("instructor", ""),
                row.get("capacity", 0),
                row.get("assigned", 0),
                row.get("remaining", 0),
            ]
        )

    summary_sheet = workbook.create_sheet("Summary")
    summary_sheet.append(["Metric", "Value"])
    for key, value in (result.get("summary") or {}).items():
        summary_sheet.append([key.replace("_", " ").title(), value])

    notes_sheet = workbook.create_sheet("Validation Warnings")
    notes_sheet.append(["Warning"])
    for warning in warnings or []:
        notes_sheet.append([warning])
    if not (warnings or []):
        notes_sheet.append(["No warnings"])

    out = io.BytesIO()
    workbook.save(out)
    return out.getvalue()
